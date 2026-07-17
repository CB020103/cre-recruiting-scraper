"""Run the top-broker search and write results directly into the Excel workbook.

Creates (or replaces) one tab per market named "<Market> Top Brokers" -
e.g. "Atlanta Top Brokers" - containing the #1 broker at each firm plus the
runners-up. Each market keeps exactly two tabs: Top Brokers (primary
reference) and RCA (supporting deal-volume data) - the old standalone
directory tabs are retired since Top Brokers replaces their job.
Formatted to be print-ready: title banner, bordered table, wrapped text,
frozen header, autofilter, landscape print setup with the header row
repeating on every printed page.

Usage:
    python update_workbook.py --workbook "Recruiting_Workbook.xlsx"
    python update_workbook.py --workbook "Recruiting_Workbook.xlsx" --market atlanta

Requirements:
    pip install openpyxl requests beautifulsoup4
"""

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break

sys.path.insert(0, str(Path(__file__).parent))
from config.markets import MARKETS, get_market
from top_brokers import process_market

COLUMNS = ["Company", "Rank", "Name", "Title", "City", "Website", "Notes"]
COLUMN_WIDTHS = [24, 6, 22, 30, 18, 42, 40]

BANNER_FILL = PatternFill("solid", fgColor="FF1F3864")
HEADER_FILL = PatternFill("solid", fgColor="FF2E5395")
BANNER_FONT = Font(name="Arial", size=13, bold=True, color="FFFFFFFF")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFFFF")
DATA_FONT = Font(name="Calibri", size=11)
THIN = Side(style="thin", color="FFB7C6E5")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MARKET_RCA_TAB = {
    "florida": "Florida RCA",
    "richmond": "Richmond RCA",
    "dmv": "DMV RCA",
    "atlanta": "Atlanta RCA",
    "boston": "Boston RCA",
    "charlotte-raleigh": "CharRal RCA",
    "nashville": "Nashville RCA",
}

TABS_TO_RETIRE = [
    "Florida", "Richmond", "DMV", "Atlanta", "Boston",
    "Charlotte + Raleigh", "Nashville", "FL Greenstreet",
]


def tab_name_for_market(display_name):
    name = f"{display_name} Top Brokers"
    for bad_char in r':\/?*[]':
        name = name.replace(bad_char, "-")
    return name[:31]


def write_market_tab(workbook, market_slug, output_rows):
    display_name = get_market(market_slug)["display_name"]
    tab_name = tab_name_for_market(display_name)

    if tab_name in workbook.sheetnames:
        del workbook[tab_name]
    ws = workbook.create_sheet(tab_name)

    n_cols = len(COLUMNS)
    last_col = get_column_letter(n_cols)

    ws.merge_cells(f"A1:{last_col}1")
    banner = ws["A1"]
    banner.value = f"{display_name} - Top Brokers by Firm"
    banner.font = BANNER_FONT
    banner.fill = BANNER_FILL
    banner.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    header_row = 2
    for col_idx, header in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER
    ws.row_dimensions[header_row].height = 20

    for row_offset, row in enumerate(output_rows):
        row_idx = header_row + 1 + row_offset
        values = [
            row["company"], row["rank"], row["name"], row["title"],
            row["city"], row["website"], row.get("notes", ""),
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = CELL_BORDER
            wrap = col_idx in (4, 6, 7)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=wrap)

    for col_idx, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = f"A{header_row + 1}"
    if output_rows:
        ws.auto_filter.ref = f"A{header_row}:{last_col}{header_row + len(output_rows)}"

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{header_row}:{header_row}"
    ws.print_options.horizontalCentered = True

    return tab_name


def retire_old_tabs(workbook):
    removed = []
    for name in TABS_TO_RETIRE:
        if name in workbook.sheetnames:
            del workbook[name]
            removed.append(name)
    return removed


def group_tabs_by_market(workbook):
    ordered_names = []
    seen = set()

    for market_slug, rca_name in MARKET_RCA_TAB.items():
        top_brokers_name = tab_name_for_market(get_market(market_slug)["display_name"])
        if top_brokers_name in workbook.sheetnames and top_brokers_name not in seen:
            ordered_names.append(top_brokers_name)
            seen.add(top_brokers_name)
        if rca_name in workbook.sheetnames and rca_name not in seen:
            ordered_names.append(rca_name)
            seen.add(rca_name)

    for name in workbook.sheetnames:
        if name not in seen:
            ordered_names.append(name)
            seen.add(name)

    workbook._sheets.sort(key=lambda ws: ordered_names.index(ws.title))


def parse_arguments():
    parser = argparse.ArgumentParser(description="Write top-broker results into the Excel workbook.")
    parser.add_argument("--workbook", required=True, help="Path to the .xlsx recruiting workbook.")
    parser.add_argument(
        "--market",
        default="all",
        choices=sorted(MARKETS) + ["all"],
        help="A specific market slug, or 'all' (default) to update every market's tab.",
    )
    return parser.parse_args()


def main():
    args = parse_arguments()
    workbook_path = Path(args.workbook)

    if not workbook_path.exists():
        raise FileNotFoundError(
            f"Could not find workbook at {workbook_path}. "
            "Pass the correct path with --workbook \"path\\to\\file.xlsx\"."
        )

    print(f"[workbook] {workbook_path}")
    workbook = load_workbook(workbook_path)

    markets_to_run = sorted(MARKETS) if args.market == "all" else [args.market]

    updated_tabs = []
    for market_slug in markets_to_run:
        output_rows = process_market(market_slug)
        tab_name = write_market_tab(workbook, market_slug, output_rows)
        updated_tabs.append(tab_name)
        print(f"[tab] wrote '{tab_name}' ({len(output_rows)} rows)")

    removed = retire_old_tabs(workbook)
    if removed:
        print(f"[retired] removed old directory tabs: {', '.join(removed)}")

    group_tabs_by_market(workbook)
    print("[order] grouped Top Brokers + RCA tabs by market")

    workbook.save(workbook_path)
    print(f"\nSaved. Updated tabs: {', '.join(updated_tabs)}")


if __name__ == "__main__":
    main()
